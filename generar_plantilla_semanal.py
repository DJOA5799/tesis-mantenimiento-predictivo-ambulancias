"""
generar_plantilla_semanal.py
Genera la plantilla semanal de programación de mantenimiento con soporte predictivo.
Toma como entrada el archivo de salida del modelo (lineamiento_priorizacion.csv)
y produce un Excel editable listo para uso operativo.

Uso:
    python generar_plantilla_semanal.py

Entrada esperada (resultados/lineamiento_priorizacion.csv):
    id_ambulancia, t0, prob_inoperatividad, nivel_riesgo,
    dias_desde_ultima_interv, downtime_total_dias_w, n_cm_w,
    disponibilidad_w, equipamiento_funcional,
    km_en_w, servicios_en_w

Salida:
    resultados/plantilla_semanal_YYYYMMDD.xlsx
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from datetime import datetime, date
import os

# ── PARÁMETROS CONFIGURABLES ──────────────────────────────────────────────────
INPUT_FILE   = "resultados/soporte_preventivo_completo.csv"
OUTPUT_DIR   = "resultados"
UMBRAL_ALTO  = 0.50
UMBRAL_MEDIO = 0.25
PLAZO_ALTO   = 3   # días hábiles
PLAZO_MEDIO  = 7   # días hábiles

# Umbrales de alerta (Anexo G)
UMBRAL_SERVICIOS   = 120
UMBRAL_KM          = 8000
UMBRAL_DIAS_SIN_PM = 45
UMBRAL_DISP        = 0.85
UMBRAL_DOWNTIME    = 7.0

# Fecha de corte para generar la plantilla.
# Usar None para tomar automáticamente el último corte disponible.
CORTE_OBJETIVO = None
# Ejemplo:
# CORTE_OBJETIVO = "2025-10-17"

# ── COLORES INSTITUCIONALES ───────────────────────────────────────────────────
C_HEADER_DARK  = "17365D"   # azul institucional
C_HEADER_MID   = "355C7D"   # azul secundario
C_HEADER_LIGHT = "FFFFFF"

C_SECTION_BG   = "F2F5F7"   # gris institucional claro
C_SUBTLE_BG    = "FAFAF6"   # fondo cálido
C_WHITE        = "FFFFFF"
C_BORDER       = "B7C1CC"
C_TEXT_DARK    = "1F2933"
C_TEXT_MUTED   = "5B6770"

# Colores de alerta, coherentes con tus nuevas figuras
C_ALTO_BG      = "F8E1DE"
C_ALTO_DARK    = "C96B63"

C_MEDIO_BG     = "FFF2BF"
C_MEDIO_DARK   = "B38F00"

C_BAJO_BG      = "E3F3EC"
C_BAJO_DARK    = "78BFA3"

C_INFO_BG      = "EAF1F7"
C_INFO_DARK    = "355C7D"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def border_thin():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="888780")
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── GENERAR DATOS DE EJEMPLO (si no existe el CSV) ────────────────────────────
def generar_datos_ejemplo():
    """Genera datos simulados del corte 12-dic-2025 para demo."""
    np.random.seed(42)
    n = 35
    ids = [f"AMB-T2-{i:03d}" for i in range(1, n+1)]
    t0 = "2025-12-12"

    probs = np.concatenate([
        np.random.uniform(0.50, 0.85, 13),
        np.random.uniform(0.25, 0.50, 20),
        np.random.uniform(0.05, 0.25,  2)
    ])
    np.random.shuffle(probs)

    def nivel(p):
        if p >= UMBRAL_ALTO:  return "ALTO"
        if p >= UMBRAL_MEDIO: return "MEDIO"
        return "BAJO"

    df = pd.DataFrame({
        "id_ambulancia":        ids,
        "t0":                   t0,
        "prob_inoperatividad":  probs.round(4),
        "nivel_riesgo":         [nivel(p) for p in probs],
        "dias_desde_ultima_interv": np.random.randint(5, 60, n),
        "downtime_total_dias_w":np.random.exponential(3, n).round(1),
        "n_cm_w":               np.random.poisson(0.8, n).astype(int),
        "disponibilidad_w":     np.random.uniform(0.80, 1.00, n).round(4),
        "equipamiento_funcional": np.random.choice([0,1], n, p=[0.05,0.95]),
        "km_en_w":              np.random.normal(2800, 600, n).round(0).astype(int),
        "servicios_en_w":       np.random.normal(45, 12, n).round(0).astype(int),
    })

    # Forzar valores extremos en alto riesgo para que activen alertas
    alto_idx = df[df["nivel_riesgo"]=="ALTO"].index[:5]
    df.loc[alto_idx[:2], "km_en_w"]  = [9200, 8700]
    df.loc[alto_idx[:2], "servicios_en_w"] = [135, 128]
    df.loc[alto_idx[2],  "downtime_total_dias_w"] = 9.4
    df.loc[alto_idx[2],  "disponibilidad_w"]      = 0.78
    df.loc[alto_idx[3],  "dias_desde_ultima_interv"] = 52

    df = df.sort_values("prob_inoperatividad", ascending=False).reset_index(drop=True)
    return df

# ── EVALUAR ALERTAS ────────────────────────────────────────────────────────────
def evaluar_alertas(row):
    alertas = []
    if row["servicios_en_w"] > UMBRAL_SERVICIOS:
        alertas.append(f"Servicios >120 ({int(row['servicios_en_w'])})")
    if row["km_en_w"] > UMBRAL_KM:
        alertas.append(f"Km >8 000 ({int(row['km_en_w'])} km)")
    if row["dias_desde_ultima_interv"] > UMBRAL_DIAS_SIN_PM:
        alertas.append(f">45 días sin PM ({int(row['dias_desde_ultima_interv'])}d)")
    if row["disponibilidad_w"] < UMBRAL_DISP:
        alertas.append(f"Disp. <85% ({row['disponibilidad_w']*100:.0f}%)")
    if row["downtime_total_dias_w"] > UMBRAL_DOWNTIME:
        alertas.append(f"Downtime >{UMBRAL_DOWNTIME}d ({row['downtime_total_dias_w']}d)")
    if row["equipamiento_funcional"] == 0:
        alertas.append("Equipo biomédico no funcional")
    return alertas

def instruccion(row, alertas):
    nivel = row["nivel_riesgo"]
    if nivel == "ALTO":
        base = "Inspección técnica general"
        if any("Km" in a for a in alertas):
            base += " + revisión tren motriz"
        if any("Downtime" in a or "Disp" in a for a in alertas):
            base += " + análisis causa raíz inoperatividad"
        if any(">45 días" in a for a in alertas):
            base += " + PM inmediato (intervalo vencido)"
        if any("biomédico" in a for a in alertas):
            base += " + reemplazo/reparación equipo biomédico"
        return base
    elif nivel == "MEDIO":
        base = "Inspección dirigida"
        if alertas:
            subsistemas = []
            if any("Km" in a or "Servicios" in a for a in alertas):
                subsistemas.append("sistema vehicular")
            if any("Downtime" in a or "Disp" in a for a in alertas):
                subsistemas.append("historial inoperatividad")
            if any(">45 días" in a for a in alertas):
                subsistemas.append("PM próximo a vencer")
            base += ": " + " + ".join(subsistemas) if subsistemas else ""
        return base
    else:
        return "Continuar operación normal. PM según cronograma."

# ── CONSTRUIR EL EXCEL ────────────────────────────────────────────────────────
def construir_excel(df):
    wb = Workbook()

    # ── HOJA 1: PLANTILLA SEMANAL ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Programación semanal"
    ws.sheet_view.showGridLines = False

    corte = df["t0"].iloc[0] if "t0" in df.columns else str(date.today())
    try:
        corte_dt = datetime.strptime(str(corte), "%Y-%m-%d")
        semana   = corte_dt.isocalendar()[1]
        anio     = corte_dt.year
        corte_str = corte_dt.strftime("%d/%m/%Y")
    except:
        corte_str = str(corte); semana = "—"; anio = 2025
    
    df["alertas"]     = df.apply(lambda r: evaluar_alertas(r), axis=1)
    df["instruccion"] = df.apply(lambda r: instruccion(r, r["alertas"]), axis=1)
    
    n_alto  = len(df[df["nivel_riesgo"]=="ALTO"])
    n_medio = len(df[df["nivel_riesgo"]=="MEDIO"])
    n_bajo  = len(df[df["nivel_riesgo"]=="BAJO"])
    total   = len(df)
    n_alertas_activas = df["alertas"].apply(len).gt(0).sum()

    # Estimación de unidades que el preventivo no hubiera programado:
    # Unidades de riesgo ALTO/MEDIO cuyo dias_desde_ultima_interv < 37 días
    
    df_no_prev = df[
        (df["nivel_riesgo"].isin(["ALTO","MEDIO"])) &
        (df["dias_desde_ultima_interv"] < 37)
    ]
    n_mejora = len(df_no_prev)

    def set_row_height(ws, row, height):
        ws.row_dimensions[row].height = height

    def merge_and_write(ws, r, c1, c2, value, fill_c, font_c, bold=True,
                        size=10, h="left", wrap=False, italic=False):
        cell = ws.cell(row=r, column=c1, value=value)
        if c1 != c2:
            ws.merge_cells(start_row=r, start_column=c1,
                           end_row=r, end_column=c2)
        cell.fill     = fill(fill_c)
        cell.font     = font(bold=bold, color=font_c, size=size, italic=italic)
        cell.alignment = align(h, "center", wrap)
        return cell
    
    def write_cell(ws, r, c, value, fill_c=C_WHITE, font_c="000000",
                   bold=False, size=9, h="left", wrap=True, italic=False):
        cell = ws.cell(row=r, column=c, value=value)
        cell.fill = fill(fill_c)
        cell.font = font(bold=bold, color=font_c, size=size, italic=italic)
        cell.alignment = align(h, "center", wrap)
        cell.border = border_thin()
        return cell

    def section_bar(ws, r, title):
        set_row_height(ws, r, 18)
        merge_and_write(
            ws, r, 1, 8, title,
            C_HEADER_DARK, C_HEADER_LIGHT,
            bold=True, size=9, h="left", wrap=False
        )
    # Column widths institucionales
    # A=Unidad, B=P, C=Riesgo, D=Alertas, E=Alertas activas,
    # F=Acción recomendada, G=Plazo, H=Responsable/Obs.
    col_w = [0, 20, 14, 14, 16, 36, 36, 18, 26]
    for i, w in enumerate(col_w[1:], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

        r = 1

    # ── CABECERA INSTITUCIONAL ────────────────────────────────────────────────
        set_row_height(ws, r, 26)
        merge_and_write(
            ws, r, 1, 8,
            "PROGRAMACIÓN SEMANAL DE MANTENIMIENTO CON SOPORTE PREDICTIVO",
            C_HEADER_DARK, C_HEADER_LIGHT,
            size=12, bold=True, h="center"
        )

        r += 1
        set_row_height(ws, r, 18)
        merge_and_write(
            ws, r, 1, 8,
            f"Ambulancias médicas urbanas Tipo II | Semana {semana} / {anio} | Corte: {corte_str}",
            C_HEADER_MID, C_HEADER_LIGHT,
            size=9, bold=False, h="center"
        )

        # Espacio
        r += 2

        # ── FICHA TÉCNICA GENERAL ────────────────────────────────────────────────
        section_bar(ws, r, "1. DATOS GENERALES DEL CORTE")
        r += 1

        ficha = [
            ("Fecha de corte", corte_str, "Flota evaluada", f"{total} unidades"),
            ("Horizonte predictivo", "14 días", "Modelo predictivo", df["modelo_base"].iloc[0] if "modelo_base" in df.columns else "Random Forest"),
            ("Umbral alto", f"P ≥ {UMBRAL_ALTO:.2f}", "Umbral medio", f"{UMBRAL_MEDIO:.2f} ≤ P < {UMBRAL_ALTO:.2f}"),
            ("Responsable", "Jefatura / Coordinación de mantenimiento", "Frecuencia sugerida", "Semanal"),
        ]

        for fila in ficha:
            set_row_height(ws, r, 24)

            # Campo 1
            write_cell(
                ws, r, 1, fila[0],
                C_SECTION_BG, C_TEXT_DARK,
                bold=True, h="center"
            )

            # Valor 1 combinado en B:C
            merge_and_write(
                ws, r, 2, 3, fila[1],
                C_WHITE, C_TEXT_DARK,
                bold=False, size=9, h="center", wrap=True
            )
            for col in range(2, 4):
                ws.cell(row=r, column=col).border = border_thin()

            # Campo 2 en D
            write_cell(
                ws, r, 4, fila[2],
                C_SECTION_BG, C_TEXT_DARK,
                bold=True, h="center"
            )

            # Valor 2 combinado en E:H
            merge_and_write(
                ws, r, 5, 8, fila[3],
                C_WHITE, C_TEXT_DARK,
                bold=False, size=9, h="center", wrap=True
            )
            for col in range(5, 9):
                ws.cell(row=r, column=col).border = border_thin()

            r += 1

        # Espacio
        r += 1

        # ── RESUMEN DE PRIORIZACIÓN ──────────────────────────────────────────────
        section_bar(ws, r, "2. RESUMEN DE PRIORIZACIÓN")
        r += 1

        resumen = [
            ("RIESGO ALTO", n_alto, f"P ≥ {UMBRAL_ALTO:.2f}", C_ALTO_BG, C_ALTO_DARK),
            ("RIESGO MEDIO", n_medio, f"{UMBRAL_MEDIO:.2f} ≤ P < {UMBRAL_ALTO:.2f}", C_MEDIO_BG, C_MEDIO_DARK),
            ("RIESGO BAJO", n_bajo, f"P < {UMBRAL_MEDIO:.2f}", C_BAJO_BG, C_BAJO_DARK),
            ("CON ALERTAS ACTIVAS", n_alertas_activas, "Alertas técnicas registradas", C_INFO_BG, C_INFO_DARK),
        ]

        # Encabezados resumen
        set_row_height(ws, r, 20)
        for idx, (label, n, criterio, bg, fg) in enumerate(resumen):
            c1 = idx * 2 + 1
            c2 = c1 + 1
            merge_and_write(ws, r, c1, c2, label, bg, fg, bold=True, size=9, h="center")
            for col in range(c1, c2 + 1):
                ws.cell(row=r, column=col).border = border_thin()

        r += 1
        set_row_height(ws, r, 24)
        for idx, (label, n, criterio, bg, fg) in enumerate(resumen):
            c1 = idx * 2 + 1
            c2 = c1 + 1
            merge_and_write(ws, r, c1, c2, str(n), C_WHITE, fg, bold=True, size=14, h="center")
            for col in range(c1, c2 + 1):
                ws.cell(row=r, column=col).border = border_thin()

        r += 1
        set_row_height(ws, r, 24)
        for idx, (label, n, criterio, bg, fg) in enumerate(resumen):
            c1 = idx * 2 + 1
            c2 = c1 + 1
            merge_and_write(ws, r, c1, c2, criterio, C_WHITE, C_TEXT_MUTED, bold=False, size=8, h="center", wrap=True)
            for col in range(c1, c2 + 1):
                ws.cell(row=r, column=col).border = border_thin()

        # Espacio
        r += 2

        # ── GUÍA DE INTERPRETACIÓN ───────────────────────────────────────────────
        section_bar(ws, r, "3. GUÍA DE INTERPRETACIÓN")
        r += 1

        guia = [
            ("Riesgo alto", f"P ≥ {UMBRAL_ALTO:.2f}", "Intervención prioritaria. Revisar la unidad en el plazo definido y validar alertas técnicas."),
            ("Riesgo medio", f"{UMBRAL_MEDIO:.2f} ≤ P < {UMBRAL_ALTO:.2f}", "Inspección programada. Confirmar condición técnica y priorizar según disponibilidad operativa."),
            ("Riesgo bajo", f"P < {UMBRAL_MEDIO:.2f}", "Seguimiento rutinario. Mantener programación preventiva institucional según corresponda."),
            ("Alertas activas", "Reglas técnicas", "Orientan el tipo de revisión: uso operativo, historial de inoperatividad, disponibilidad o equipamiento biomédico."),
        ]

        # Headers guía
        set_row_height(ws, r, 20)
        headers_guia = ["Nivel / criterio", "Rango o fuente", "Interpretación operativa"]
        write_cell(ws, r, 1, headers_guia[0], C_HEADER_MID, C_HEADER_LIGHT, bold=True, h="center")
        write_cell(ws, r, 2, headers_guia[1], C_HEADER_MID, C_HEADER_LIGHT, bold=True, h="center")
        merge_and_write(ws, r, 3, 8, headers_guia[2], C_HEADER_MID, C_HEADER_LIGHT, bold=True, size=9, h="center", wrap=True)
        for col in range(3, 9):
            ws.cell(row=r, column=col).border = border_thin()

        r += 1
        for nivel_txt, rango, interpretacion in guia:
            set_row_height(ws, r, 28)
            if "alto" in nivel_txt.lower():
                bg, fg = C_ALTO_BG, C_ALTO_DARK
            elif "medio" in nivel_txt.lower():
                bg, fg = C_MEDIO_BG, C_MEDIO_DARK
            elif "bajo" in nivel_txt.lower():
                bg, fg = C_BAJO_BG, C_BAJO_DARK
            else:
                bg, fg = C_INFO_BG, C_INFO_DARK

            write_cell(ws, r, 1, nivel_txt, bg, fg, bold=True, h="center")
            write_cell(ws, r, 2, rango, C_WHITE, C_TEXT_DARK, h="center")
            merge_and_write(ws, r, 3, 8, interpretacion, C_WHITE, C_TEXT_DARK, bold=False, size=8, h="left", wrap=True)
            for col in range(3, 9):
                ws.cell(row=r, column=col).border = border_thin()
            r += 1
    
    # ── TABLA PRINCIPAL ───────────────────────────────────────────────────────
        r += 1
        section_bar(ws, r, "4. PROGRAMACIÓN SEMANAL PRIORIZADA")
        r += 1

        set_row_height(ws, r, 28)
        headers = [
            (1, "Unidad"),
            (2, "P(inop.)"),
            (3, "Riesgo"),
            (4, "N.° alertas"),
            (5, "Alertas activas"),
            (6, "Acción recomendada"),
            (7, "Plazo sugerido"),
            (8, "Observaciones")
        ]

        for c, h in headers:
            cell = ws.cell(row=r, column=c, value=h)
            cell.fill = fill(C_HEADER_DARK)
            cell.font = font(bold=True, color=C_HEADER_LIGHT, size=9)
            cell.alignment = align("center", "center", True)
            cell.border = border_thin()

        # Orden institucional: alto, medio, bajo; dentro de cada grupo, mayor probabilidad primero
        orden_riesgo = {"ALTO": 1, "MEDIO": 2, "BAJO": 3}
        df_tabla = df.copy()
        df_tabla["orden_riesgo"] = df_tabla["nivel_riesgo"].map(orden_riesgo)
        df_tabla = df_tabla.sort_values(
            ["orden_riesgo", "prob_inoperatividad"],
            ascending=[True, False]
        ).reset_index(drop=True)

        for _, row_data in df_tabla.iterrows():
            r += 1
            set_row_height(ws, r, 34)

            nivel = row_data["nivel_riesgo"]
            alertas = row_data["alertas"]
            n_alertas = len(alertas)
            alerta_str = "\n".join(alertas) if alertas else "—"

            if nivel == "ALTO":
                row_bg, fg = "FFF5F4", C_ALTO_DARK
                plazo = f"≤ {PLAZO_ALTO} días hábiles"
            elif nivel == "MEDIO":
                row_bg, fg = "FFFAE8", C_MEDIO_DARK
                plazo = f"≤ {PLAZO_MEDIO} días hábiles"
            else:
                row_bg, fg = "F3FAF6", C_BAJO_DARK
                plazo = "Según cronograma"

            vals = [
                row_data["id_ambulancia"],
                round(float(row_data["prob_inoperatividad"]), 4),
                nivel,
                n_alertas,
                alerta_str,
                row_data["instruccion"],
                plazo,
                ""
            ]

            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.fill = fill(row_bg)
                cell.border = border_thin()
                cell.alignment = align("center" if c in [2,3,4,7] else "left", "center", True)
                cell.font = font(
                    bold=(c in [1, 3]),
                    color=fg if c == 3 else C_TEXT_DARK,
                    size=8
                )
                if c == 2:
                    cell.number_format = "0.0000"

    # ── CONTROL Y SEGUIMIENTO ─────────────────────────────────────────────────
        r += 3
        section_bar(ws, r, "5. CONTROL Y SEGUIMIENTO SEMANAL")
        r += 1

        control_headers = [
            "Total alto", "Total medio", "Total bajo", "Con alertas",
            "Atendidas", "Pendientes", "Reprogramadas", "Cierre"
        ]

        control_values = [
            n_alto, n_medio, n_bajo, n_alertas_activas,
            "", "", "", ""
        ]

        set_row_height(ws, r, 22)
        for c, h in enumerate(control_headers, 1):
            cell = ws.cell(row=r, column=c, value=h)
            cell.fill = fill(C_SECTION_BG)
            cell.font = font(bold=True, color=C_TEXT_DARK, size=8)
            cell.alignment = align("center", "center", True)
            cell.border = border_thin()

        r += 1
        set_row_height(ws, r, 26)
        for c, v in enumerate(control_values, 1):
            if c == 1:
                bg, fg = C_ALTO_BG, C_ALTO_DARK
            elif c == 2:
                bg, fg = C_MEDIO_BG, C_MEDIO_DARK
            elif c == 3:
                bg, fg = C_BAJO_BG, C_BAJO_DARK
            elif c == 4:
                bg, fg = C_INFO_BG, C_INFO_DARK
            else:
                bg, fg = C_WHITE, C_TEXT_DARK

            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill(bg)
            cell.font = font(bold=True, color=fg, size=11 if c <= 4 else 9)
            cell.alignment = align("center", "center", True)
            cell.border = border_thin()

        r += 1
        set_row_height(ws, r, 28)
        merge_and_write(
            ws, r, 1, 8,
            "Observaciones generales: Las acciones programadas deben validarse con la disponibilidad operativa, "
            "recursos técnicos, repuestos y criterio de la jefatura de mantenimiento.",
            C_WHITE, C_TEXT_DARK,
            bold=False, size=8, h="left", wrap=True
        )
        for col in range(1, 9):
            ws.cell(row=r, column=col).border = border_thin()

    # ── FIRMAS / VALIDACIÓN ───────────────────────────────────────────────────
        r += 3
        section_bar(ws, r, "6. VALIDACIÓN DEL DOCUMENTO")
        r += 1

        set_row_height(ws, r, 20)
        merge_and_write(ws, r, 1, 2, "Elaborado por", C_SECTION_BG, C_TEXT_DARK, bold=True, size=8, h="center")
        merge_and_write(ws, r, 3, 5, "Revisado por", C_SECTION_BG, C_TEXT_DARK, bold=True, size=8, h="center")
        merge_and_write(ws, r, 6, 8, "Aprobado por", C_SECTION_BG, C_TEXT_DARK, bold=True, size=8, h="center")

        for col in range(1, 9):
            ws.cell(row=r, column=col).border = border_thin()

        r += 1
        set_row_height(ws, r, 42)
        merge_and_write(ws, r, 1, 2, f"Nombre / firma:\nFecha: {corte_str}", C_WHITE, C_TEXT_DARK, bold=False, size=8, h="left", wrap=True)
        merge_and_write(ws, r, 3, 5, "Nombre / firma:\nFecha: ___/___/____", C_WHITE, C_TEXT_DARK, bold=False, size=8, h="left", wrap=True)
        merge_and_write(ws, r, 6, 8, "Nombre / firma:\nFecha: ___/___/____", C_WHITE, C_TEXT_DARK, bold=False, size=8, h="left", wrap=True)

        for col in range(1, 9):
            ws.cell(row=r, column=col).border = border_thin()

        modelo_base = df["modelo_base"].iloc[0] if "modelo_base" in df.columns else "Random Forest"

        # ── NOTA PIE ─────────────────────────────────────────────────────────────
        r += 2
        set_row_height(ws, r, 34)
        merge_and_write(
            ws, r, 1, 8,
            f"Fuente: Elaboración propia. Plantilla generada a partir de probabilidades estimadas por el modelo predictivo {modelo_base}. "
            "Documento metodológico elaborado con datos simulados para fines de investigación. "
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}.",
            C_SECTION_BG, C_TEXT_MUTED,
            bold=False, size=8, h="left", wrap=True, italic=True
        )

    # ── HOJA 2: DATOS COMPLETOS ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Datos completos")
    ws2.sheet_view.showGridLines = True

    export_cols = [
        "id_ambulancia","t0","prob_inoperatividad","nivel_riesgo",
        "dias_desde_ultima_interv","downtime_total_dias_w","n_cm_w",
        "disponibilidad_w","equipamiento_funcional","km_en_w","servicios_en_w"
    ]
    df_export = df[export_cols].copy()
    df_export["alertas_activas"] = df["alertas"].apply(lambda x: " | ".join(x) if x else "—")
    df_export["instruccion_taller"] = df["instruccion"]

    headers2 = list(df_export.columns)
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = fill(C_HEADER_DARK)
        cell.font = font(bold=True, color=C_HEADER_LIGHT, size=9)
        cell.alignment = align("center","center",True)

    for ri, (_, row_data) in enumerate(df_export.iterrows(), start=2):
        nivel = row_data["nivel_riesgo"]
        row_bg = (C_ALTO_BG if nivel=="ALTO" else
                  C_MEDIO_BG if nivel=="MEDIO" else C_BAJO_BG)

        for c, v in enumerate(row_data, 1):
            cell = ws2.cell(row=ri, column=c, value=v)
            cell.fill = fill(row_bg)
            cell.font = font(size=9, color=C_TEXT_DARK)
            cell.alignment = align("center", "center", True)
            cell.border = border_thin()

    for c in range(1, len(headers2)+1):
        ws2.column_dimensions[get_column_letter(c)].width = 18
    ws2.freeze_panes = "A2"

    # ── GUARDAR ───────────────────────────────────────────────────────────────
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    try:
        corte_fmt = datetime.strptime(str(corte), "%Y-%m-%d").strftime("%Y%m%d")
    except:
        corte_fmt = datetime.now().strftime("%Y%m%d")

    out_path = os.path.join(OUTPUT_DIR,
                            f"plantilla_semanal_{corte_fmt}.xlsx")
    wb.save(out_path)
    return out_path

# ── MAIN ──────────────────────────────────────────────────────────────────────
if os.path.exists(INPUT_FILE):
    df = pd.read_csv(INPUT_FILE)
    print(f"  Archivo cargado: {INPUT_FILE} ({len(df)} filas)")

    if "t0" in df.columns:
        df["t0"] = pd.to_datetime(df["t0"])

        if CORTE_OBJETIVO is not None:
            corte_objetivo_dt = pd.to_datetime(CORTE_OBJETIVO)
            
            if corte_objetivo_dt not in df["t0"].unique():
                cortes_disponibles = sorted(df["t0"].dt.strftime("%Y-%m-%d").unique())
                raise ValueError(
                    f"El corte {CORTE_OBJETIVO} no existe en el archivo. "
                    f"Cortes disponibles: {cortes_disponibles[:5]} ... {cortes_disponibles[-5:]}"
                )

            corte_seleccionado = corte_objetivo_dt
            print(f"  Corte seleccionado manualmente: {corte_seleccionado.date()}")

        else:
            corte_seleccionado = df["t0"].max()
            print(f"  Corte seleccionado automáticamente: {corte_seleccionado.date()}")

        df = df[df["t0"] == corte_seleccionado].copy().reset_index(drop=True)
        df["t0"] = df["t0"].dt.strftime("%Y-%m-%d")
        print(f"  Unidades incluidas en plantilla: {len(df)}")

    else:
        print(f"  Archivo '{INPUT_FILE}' no encontrado.")
        print("  Generando datos de ejemplo para demostración...")
        df = generar_datos_ejemplo()

    print("Generando Excel...")
    path = construir_excel(df)
    print(f"\n✓ Plantilla generada: {path}")
    print(f"  Unidades ALTO:  {len(df[df['nivel_riesgo']=='ALTO'])}")
    print(f"  Unidades MEDIO: {len(df[df['nivel_riesgo']=='MEDIO'])}")
    print(f"  Unidades BAJO:  {len(df[df['nivel_riesgo']=='BAJO'])}")
